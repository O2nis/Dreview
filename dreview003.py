import re
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import datetime as dt
import numpy as np
import seaborn as sns
from cycler import cycler
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO

plt.rcParams.update({'font.size': 8})

def robust_parse_date(d):
    """
    Enhanced date parser that handles multiple formats and converts to standard format
    Returns: 
        - pd.Timestamp for valid dates
        - pd.NaT for invalid dates
    """
    if pd.isna(d) or d in ['', '########', '0-Jan-00', '00-Jan-00', 'NaN', 'NaT']:
        return pd.NaT
    
    if isinstance(d, pd.Timestamp):
        return d
    
    if isinstance(d, dt.datetime):
        return pd.Timestamp(d)
    
    if isinstance(d, str):
        # Normalize case for month abbreviations
        d = d.upper()
        if re.match(r'^\d{2}-[A-Z]{3}-\d{2}$', d):
            try:
                parsed = pd.to_datetime(d, format='%d-%b-%y', errors='coerce')
                if pd.notna(parsed):
                    return parsed
            except Exception as e:
                st.warning(f"Failed to parse date '{d}' with format '%d-%b-%y': {str(e)}")
        
        date_formats = [
            '%d-%b-%y', '%d-%B-%y', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d',
            '%b %d, %Y', '%B %d, %Y', '%d.%m.%Y', '%Y%m%d', '%m-%d-%Y', '%d %b %Y'
        ]
        
        for fmt in date_formats:
            try:
                parsed = pd.to_datetime(d, format=fmt, errors='coerce')
                if pd.notna(parsed):
                    return parsed
            except Exception as e:
                pass
        
        try:
            parsed = pd.to_datetime(d, errors='coerce')
            if pd.notna(parsed):
                return parsed
        except Exception as e:
            pass
        
        st.warning(f"Could not parse date: '{d}'")
    
    return pd.NaT

def is_rev_col(col_name):
    """Check if column name matches a revision-like pattern (e.g., Rev0, Revision 1)."""
    col_norm = normalize_header(col_name)
    return bool(re.search(r'rev(ision)?\s*\d+', col_norm, re.IGNORECASE))

def is_review_col(col_name):
    """Check if column name matches a review-like pattern (e.g., Reviewed0, Review Date)."""
    col_norm = normalize_header(col_name)
    return bool(re.search(r'review(ed)?(\s*date)?\s*\d*', col_norm, re.IGNORECASE))

def normalize_header(h):
    """Normalize header by removing extra spaces and converting to lowercase."""
    return ' '.join(str(h).lower().split())

def get_final_milestone(row):
    issued = pd.notna(row["Issued by EPC"])
    reviewed = pd.notna(row["Review By OE"])
    replied = pd.notna(row["Reply By EPC"])
    flag = int(row.get("Flag", 0)) == 1

    if replied and reviewed and issued and flag:
        return "Finalized"
    elif replied and reviewed and issued:
        return "Reply By EPC"
    elif reviewed and issued:
        return "Review By OE"
    elif issued:
        return "Issued by EPC"
    else:
        return "NO ISSUANCE"

def main():
    st.set_page_config(page_title="S-Curve Analysis", layout="wide")

    # Create tabs
    tab1, tab2 = st.tabs(["S-Curve Analysis", "Review Timeline"])

    with tab1:
        # ----------------------------------------------------------------
        # A) TEMPLATE CSV DOWNLOAD (with first 3 rows + headers)
        # ----------------------------------------------------------------
        TEMPLATE_CONTENT = """ID,Discipline,Area,Document Title,Project Indentifer,Originator,Document Number,Document Type ,Counter ,Revision,Area code,Disc,Category,Transmittal Code,Comment Sheet OE,Comment Sheet EPC,Schedule [Days],Issued by EPC,Issuance Expected,Review By OE,Expected review,Reply By EPC,Final Issuance Expected,Review1,ReSub1,Review2,ReSub2,Review3,ReSub3,Review4,ReSub4,Review5,ReSub5,Man Hours ,Status,CS rev,Flag
1,General,General,Overall site layout,KFE,SC,0001,MA,00,A,GEN,GN,DRG,"KFE-SC-MOEM-T-0052-AO-PV System Analysis Report, Project Quality Management Plan and TCO.",MOEM-TCO-CI-0017-Rev_00_FI,MOEM-TCO-CI-0017-Rev_00_CE,10,4-Apr-25,,4-Apr-25, , , ,,,,,,,,10,CO,1,0
2,General,General,Overall Single Line Diagram (PV plant + interconnection facilities),KFE,SC,0002,MA,00,A,GEN,GN,DRG,,,,10,18-Mar-25,,18-Mar-25, , , ,,,,,,,,10,CO,1,0
3,PV,General,PVsyst yield estimates,KFE,SC,0003,MA,00,A,GEN,PV,DRG,,,,10,,,, , , ,,,,,,,,10,FN,0,0
"""

        st.subheader("Download CSV Template")
        st.download_button(
            label="Download Template CSV",
            data=TEMPLATE_CONTENT.encode("utf-8"),
            file_name="EDDR_template.csv",
            mime="text/csv"
        )

        # --------------------------
        # 1) SIDEBAR INPUTS
        # --------------------------
        st.sidebar.header("Configuration")
        CSV_INPUT_PATH = st.sidebar.file_uploader("Upload Input File", type=["csv", "xlsx", "xls"])
        INITIAL_DATE = st.sidebar.date_input("Initial Date for Expected Calculations", value=pd.to_datetime("2024-08-01"))
        IFR_WEIGHT = st.sidebar.number_input("Issued By EPC Weight", value=0.40, step=0.05)
        IFA_WEIGHT = st.sidebar.number_input("Review By OE Weight", value=0.30, step=0.05)
        IFT_WEIGHT = st.sidebar.number_input("Reply By EPC Weight", value=0.30, step=0.05)
        RECOVERY_FACTOR = st.sidebar.number_input("Recovery Factor", value=0.75, step=0.05)
        IFA_DELTA_DAYS = st.sidebar.number_input("Days to add for Expected Review", value=10, step=1)
        IFT_DELTA_DAYS = st.sidebar.number_input("Days to add for Final Issuance Expected", value=5, step=1)

        IGNORE_STATUS = st.sidebar.text_input("Status to Ignore (comma-separated, case-sensitive, leave blank to include all)", value="")
        PERCENTAGE_VIEW = st.sidebar.checkbox("Show values as percentage of total", value=False)
        INCLUDE_COMPLETED = st.sidebar.checkbox("Include Completed Documents (Flag=1) in Delays Table", value=True)

        st.sidebar.markdown("---")
        st.sidebar.markdown("### Visualization Settings")
        seaborn_style = st.sidebar.selectbox(
            "Seaborn Style",
            ["darkgrid", "whitegrid", "dark", "white", "ticks"],
            index=1
        )
        seaborn_context = st.sidebar.selectbox(
            "Seaborn Context",
            ["paper", "notebook", "talk", "poster"],
            index=1
        )
        color_scheme = st.sidebar.selectbox(
            "Color Scheme (Bar/Donut/Pie Charts)",
            ["Standard", "Shades of Blue", "Shades of Green", "Seaborn Palette"],
            index=1
        )
        seaborn_palette = st.sidebar.selectbox(
            "Seaborn Palette (if Seaborn Palette selected)",
            ["deep", "muted", "bright", "pastel", "dark", "colorblind", "Set1", "Set2", "Set3"],
            index=0
        )
        font_scale = st.sidebar.slider("Font Scale", min_value=0.5, max_value=2.0, value=1.0, step=0.1)
        show_grid = st.sidebar.checkbox("Show Grid Lines", value=True)

        sns.set_style(seaborn_style)
        sns.set_context(seaborn_context, font_scale=font_scale)

        st.sidebar.markdown("### S-Curve Color Scheme")
        actual_color = st.sidebar.color_picker("Actual Progress Color", "#1f77b4")
        expected_color = st.sidebar.color_picker("Expected Progress Color", "#ff7f0e")
        projected_color = st.sidebar.color_picker("Projected Recovery Color", "#2ca02c")
        today_color = st.sidebar.color_picker("Today Line Color", "#000000")
        end_date_color = st.sidebar.color_picker("End Date Line Color", "#d62728")

        if CSV_INPUT_PATH is None:
            st.warning("Please upload your input CSV or Excel file or download the template above.")
            return

        # --------------------------
        # 2) LOAD CSV OR EXCEL & PREP DATA WITH ROBUST DATE PARSING
        # --------------------------
        file_extension = CSV_INPUT_PATH.name.split('.')[-1].lower()
        if file_extension in ['xlsx', 'xls']:
            df = pd.read_excel(CSV_INPUT_PATH)
        else:
            df = pd.read_csv(CSV_INPUT_PATH)

        # Define the full column list including the 10 new columns
        expected_columns = [
            "ID", "Discipline", "Area", "Document Title", "Project Indentifer", "Originator",
            "Document Number", "Document Type ", "Counter ", "Revision", "Area code",
            "Disc", "Category", "Transmittal Code", "Comment Sheet OE", "Comment Sheet EPC",
            "Schedule [Days]", "Issued by EPC", "Issuance Expected", "Review By OE",
            "Expected review", "Reply By EPC", "Final Issuance Expected",
            "Review1", "ReSub1", "Review2", "ReSub2", "Review3", "ReSub3",
            "Review4", "ReSub4", "Review5", "ReSub5",
            "Man Hours ", "Status", "CS rev", "Flag"
        ]
        df.columns = expected_columns[:len(df.columns)]  # Assign only up to the number of columns present

        if IGNORE_STATUS.strip():
            statuses_to_exclude = [s.strip() for s in IGNORE_STATUS.split(',') if s.strip()]
            initial_len = len(df)
            df = df[~df["Status"].isin(statuses_to_exclude)]
            filtered_len = len(df)
            if filtered_len < initial_len:
                st.info(f"Filtered out {initial_len - filtered_len} rows with Status in: {', '.join(statuses_to_exclude)}")
            if filtered_len == 0:
                st.error(f"All rows have Status in exclusion list. No data remains after filtering.")
                return

        date_columns = ["Issued by EPC", "Review By OE", "Reply By EPC", "Issuance Expected", "Expected review", "Final Issuance Expected"]
        for col in date_columns:
            df[col] = df[col].apply(robust_parse_date)
            na_count = df[col].isna().sum()
            if na_count > 0:
                st.warning(f"Column '{col}' has {na_count} dates that couldn't be parsed")

        df["Schedule [Days]"] = pd.to_numeric(df["Schedule [Days]"], errors="coerce").fillna(0)
        df["Man Hours "] = pd.to_numeric(df["Man Hours "], errors="coerce").fillna(0)
        df["Flag"] = pd.to_numeric(df["Flag"], errors="coerce").fillna(0)

        df["Issuance Expected"] = pd.Timestamp(INITIAL_DATE) + pd.to_timedelta(df["Schedule [Days]"], unit="D")
        df["Expected review"] = df["Issuance Expected"] + dt.timedelta(days=IFA_DELTA_DAYS)
        df["Final Issuance Expected"] = df["Expected review"] + dt.timedelta(days=IFT_DELTA_DAYS)
        df["Final Issuance Expected"] = pd.to_datetime(df["Final Issuance Expected"], errors='coerce')

        ift_expected_max = df["Final Issuance Expected"].dropna().max()
        if pd.isna(ift_expected_max):
            st.warning("No valid Final Issuance Expected dates found. Checking other date columns.")
            date_cols = ["Issuance Expected", "Expected review", "Final Issuance Expected", 
                        "Issued by EPC", "Review By OE", "Reply By EPC"]
            all_dates = df[date_cols].values.ravel()
            valid_dates = pd.Series(all_dates).dropna()
            if valid_dates.empty:
                st.error("No valid milestone dates found in any date columns. Cannot generate S-Curve.")
                return
            ift_expected_max = valid_dates.max()
        else:
            date_cols = ["Issuance Expected", "Expected review", "Final Issuance Expected", 
                        "Issued by EPC", "Review By OE", "Reply By EPC"]
            all_dates = df[date_cols].values.ravel()
            valid_dates = pd.Series(all_dates).dropna()

        if valid_dates.empty:
            st.error("No valid dates found in any milestone columns. Cannot proceed with S-Curve plotting.")
            return

        start_date = valid_dates.min()
        today_date = pd.Timestamp.today().normalize()  # Uses actual current date
        total_mh = df["Man Hours "].sum()

        # Validate timeline
        if start_date > today_date:
            st.warning(f"Start date ({start_date.strftime('%d-%b-%Y')}) is after today ({today_date.strftime('%d-%b-%Y')}). Using single point timeline.")
            actual_timeline = [today_date]
        else:
            actual_timeline = pd.date_range(start=start_date, end=today_date, freq='W')
            if len(actual_timeline) == 0:
                st.warning("Actual timeline is empty. Using single point at today.")
                actual_timeline = [today_date]

        if start_date > ift_expected_max:
            st.warning(f"Start date ({start_date.strftime('%d-%b-%Y')}) is after max expected date ({ift_expected_max.strftime('%d-%b-%Y')}). Using single point timeline.")
            expected_timeline = [ift_expected_max]
        else:
            expected_timeline = pd.date_range(start=start_date, end=ift_expected_max, freq='W')
            if len(expected_timeline) == 0:
                st.warning("Expected timeline is empty. Using single point at max expected date.")
                expected_timeline = [ift_expected_max]

        # --------------------------
        # 3) BUILD ACTUAL AND EXPECTED CUMULATIVE VALUES
        # --------------------------
        actual_cum = []
        issuance_cums = []
        review_cums = []
        final_cums = []
        last_actual_value = 0.0
        last_progress_date = start_date
        
        for current_date in actual_timeline:
            iss_sum = 0.0
            rev_sum = 0.0
            fin_sum = 0.0
            has_progress = False
            for _, row in df.iterrows():
                mh = row["Man Hours "]
                if pd.notna(row["Issued by EPC"]) and row["Issued by EPC"] <= current_date:
                    iss_sum += mh * IFR_WEIGHT
                if pd.notna(row["Review By OE"]) and row["Review By OE"] <= current_date:
                    rev_sum += mh * IFA_WEIGHT
                if pd.notna(row["Reply By EPC"]) and row["Reply By EPC"] <= current_date and row["Flag"] == 1:
                    fin_sum += mh * IFT_WEIGHT
            a_sum = iss_sum + rev_sum + fin_sum
            if a_sum > last_actual_value:
                has_progress = True
            if has_progress:
                last_actual_value = a_sum
                last_progress_date = current_date
                actual_cum.append(a_sum)
            else:
                actual_cum.append(last_actual_value)
            issuance_cums.append(iss_sum)
            review_cums.append(rev_sum)
            final_cums.append(fin_sum)
        
        expected_cum = []
        last_expected_value = 0.0
        last_expected_progress_date = start_date
        
        for current_date in expected_timeline:
            e_sum = 0.0
            has_progress = False
            for _, row in df.iterrows():
                mh = row["Man Hours "]
                e_val = 0.0
                if pd.notna(row["Issuance Expected"]) and row["Issuance Expected"] <= current_date:
                    e_val += IFR_WEIGHT
                if pd.notna(row["Expected review"]) and row["Expected review"] <= current_date:
                    e_val += IFA_WEIGHT
                if pd.notna(row["Final Issuance Expected"]) and row["Final Issuance Expected"] <= current_date:
                    e_val += IFT_WEIGHT
                if e_val > 0:
                    has_progress = True
                e_sum += mh * e_val
            if has_progress:
                last_expected_value = e_sum
                last_expected_progress_date = current_date
                expected_cum.append(e_sum)
            else:
                expected_cum.append(last_expected_value)
        
        if not actual_cum or not expected_cum:
            st.error("No cumulative progress data generated. Check input data for valid dates and man-hours.")
            return

        final_actual = actual_cum[-1]
        final_expected = expected_cum[-1]

        # Extend timelines if necessary
        if last_progress_date < today_date:
            actual_timeline = list(actual_timeline) + [today_date]
            actual_cum = actual_cum + [last_actual_value]
            issuance_cums = issuance_cums + [issuance_cums[-1]]
            review_cums = review_cums + [review_cums[-1]]
            final_cums = final_cums + [final_cums[-1]]
        
        if pd.notna(ift_expected_max) and last_expected_progress_date < ift_expected_max:
            expected_timeline = list(expected_timeline) + [ift_expected_max]
            expected_cum = expected_cum + [last_expected_value]

        # --------------------------
        # 4) PROJECTED RECOVERY LINE
        # --------------------------
        if today_date <= actual_timeline[0]:
            today_idx = 0
        elif today_date >= actual_timeline[-1]:
            today_idx = len(actual_timeline) - 1
        else:
            today_idx = np.searchsorted(actual_timeline, today_date, side="right") - 1

        actual_today = actual_cum[today_idx]
        if today_date <= expected_timeline[0]:
            expected_today_idx = 0
        elif today_date >= expected_timeline[-1]:
            expected_today_idx = len(expected_timeline) - 1
        else:
            expected_today_idx = np.searchsorted(expected_timeline, today_date, side="right") - 1
        expected_today = expected_cum[expected_today_idx]

        gap_hrs = final_expected - actual_today
        projected_timeline = []
        projected_cumulative = []
        recovery_end_date = None

        if gap_hrs > 0:
            total_days_span = (ift_expected_max - start_date).days
            project_months = total_days_span / 30.4
            delay_fraction = gap_hrs / final_expected
            T_recover_months = project_months * delay_fraction * RECOVERY_FACTOR
            T_recover_weeks = T_recover_months * (30.4 / 7.0)
            if T_recover_weeks < 1:
                T_recover_weeks = 1
            slope_new = gap_hrs / T_recover_weeks
            last_date = today_date
            cum_val = actual_today
            steps = int(T_recover_weeks) + 2
            for _ in range(steps):
                projected_timeline.append(last_date)
                projected_cumulative.append(cum_val)
                if cum_val >= final_expected:
                    break
                last_date = last_date + dt.timedelta(weeks=1)
                cum_val = min(final_expected, cum_val + slope_new)
            recovery_end_date = last_date

        # --------------------------
        # 5) S-CURVE
        # --------------------------
        st.subheader("S-Curve with Delay Recovery")
        y_actual = [x/total_mh*100 for x in actual_cum] if PERCENTAGE_VIEW else actual_cum
        y_expected = [x/total_mh*100 for x in expected_cum] if PERCENTAGE_VIEW else expected_cum
        y_projected = [x/total_mh*100 for x in projected_cumulative] if PERCENTAGE_VIEW and projected_cumulative else projected_cumulative
        y_label = "Cumulative % of Total Works" if PERCENTAGE_VIEW else "Cumulative Man-Hours"

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(actual_timeline, y_actual, label="Actual Progress", color=actual_color, linewidth=2)
        ax.plot(expected_timeline, y_expected, label="Expected Progress", color=expected_color, linewidth=2)

        if last_progress_date < today_date:
            ax.hlines(y=y_actual[-1], xmin=last_progress_date, xmax=today_date, 
                     color=actual_color, linestyle='-', linewidth=2)
        
        if pd.notna(ift_expected_max) and last_expected_progress_date < ift_expected_max:
            ax.hlines(y=y_expected[-1], xmin=last_expected_progress_date, xmax=ift_expected_max, 
                     color=expected_color, linestyle='-', linewidth=2)

        if projected_timeline:
            ax.plot(
                projected_timeline, y_projected, linestyle=":", label="Projected (Recovery Factor)",
                color=projected_color, linewidth=3
            )

        ax.set_title("S-Curve with Delay Recovery", fontsize=12)
        ax.set_xlabel("Date", fontsize=10)
        ax.set_ylabel(y_label, fontsize=10)
        if show_grid:
            ax.grid(True)
        ax.legend(fontsize=9)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d-%b-%Y"))
        plt.xticks(rotation=45)
        plt.tight_layout()

        ax.axvline(today_date, color=today_color, linestyle="--", linewidth=1.5, label="Today")
        ax.annotate(
            f"Today\n{today_date.strftime('%d-%b-%Y')}",
            xy=(today_date, (y_expected[-1] if PERCENTAGE_VIEW else final_expected) * 0.1),
            xytext=(10, 10), textcoords="offset points", color=today_color,
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="none", alpha=0.7), fontsize=8
        )

        if pd.notna(ift_expected_max):
            ax.axvline(ift_expected_max, linestyle="--", linewidth=1.5, color=end_date_color)
            ax.annotate(
                f"Original End\n{ift_expected_max.strftime('%d-%b-%Y')}",
                xy=(ift_expected_max, (y_expected[-1] if PERCENTAGE_VIEW else final_expected) * 0.2),
                xytext=(-100, 10), textcoords="offset points", color=end_date_color,
                bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="none", alpha=0.7),
                fontsize=8, arrowprops=dict(arrowstyle="->", color=end_date_color)
            )

        if recovery_end_date:
            ax.axvline(recovery_end_date, linestyle="--", linewidth=1.5, color=end_date_color)
            ax.annotate(
                f"Recovery End\n{recovery_end_date.strftime('%d-%b-%Y')}",
                xy=(recovery_end_date, (y_expected[-1] if PERCENTAGE_VIEW else final_expected) * 0.3),
                xytext=(10,10), textcoords="offset points", color=end_date_color,
                bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="none", alpha=0.7),
                fontsize=8, arrowprops=dict(arrowstyle="->", color=end_date_color)
            )

        delay_today = expected_today - actual_today
        delay_pct = (delay_today / final_expected * 100) if final_expected > 0 else 0
        
        # Always show Actual Progress as a percentage (even if the chart is in MH)
        actual_pct = (y_actual[today_idx] if PERCENTAGE_VIEW
                      else ((actual_today / total_mh * 100) if total_mh > 0 else 0))
        
        delay_text = (
            f"Actual Progress: {actual_pct:.1f}%\n"
            + (f"Current Delay: {delay_pct:.1f}%"
               if PERCENTAGE_VIEW
               else f"Current Delay: {delay_today:,.1f} MH\n({delay_pct:.1f}%)")
        )

        ax.annotate(
            delay_text, xy=(today_date, (y_actual[today_idx] + y_expected[expected_today_idx])/2),
            xytext=(10, -10), textcoords="offset points", color=today_color,
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="none", alpha=0.7),
            arrowprops=dict(arrowstyle="->", color=today_color), ha="left", fontsize=8
        )

        st.pyplot(fig)

        # --------------------------
        # 6) COLOR SCHEME FOR OTHER CHARTS
        # --------------------------
        standard_cycler = plt.rcParamsDefault['axes.prop_cycle']
        blue_cycler = cycler(color=["#cce5ff", "#99ccff", "#66b2ff", "#3399ff", "#007fff"])
        green_cycler = cycler(color=["#ccffcc", "#99ff99", "#66ff66", "#33cc33", "#009900"])
        if color_scheme == "Standard":
            plt.rc("axes", prop_cycle=standard_cycler)
            stack_colors = ["#1f77b4", "#ff7f0e", "#2ca02c"]  # Match S-Curve colors
        elif color_scheme == "Shades of Blue":
            plt.rc("axes", prop_cycle=blue_cycler)
            stack_colors = ["#cce5ff", "#66b2ff", "#007fff"]  # Shades of blue
        elif color_scheme == "Shades of Green":
            plt.rc("axes", prop_cycle=green_cycler)
            stack_colors = ["#ccffcc", "#66ff66", "#009900"]  # Shades of green
        else:
            palette_colors = sns.color_palette(seaborn_palette, n_colors=10)
            plt.rc("axes", prop_cycle=cycler(color=palette_colors))
            stack_colors = palette_colors[:3]  # Use first three colors from palette

        # --------------------------
        # 7) ACTUAL vs EXPECTED HOURS BY DISCIPLINE
        # --------------------------
        end_date = max(today_date, ift_expected_max)
        final_actual_progress = []
        final_expected_progress = []
        for _, row in df.iterrows():
            a_prog = 0.0
            if pd.notna(row["Issued by EPC"]) and row["Issued by EPC"] <= end_date:
                a_prog += IFR_WEIGHT
            if pd.notna(row["Review By OE"]) and row["Review By OE"] <= end_date:
                a_prog += IFA_WEIGHT
            if pd.notna(row["Reply By EPC"]) and row["Reply By EPC"] <= end_date and row["Flag"] == 1:
                a_prog += IFT_WEIGHT
            e_prog = 0.0
            if pd.notna(row["Issuance Expected"]) and row["Issuance Expected"] <= end_date:
                e_prog += IFR_WEIGHT
            if pd.notna(row["Expected review"]) and row["Expected review"] <= end_date:
                e_prog += IFA_WEIGHT
            if pd.notna(row["Final Issuance Expected"]) and row["Final Issuance Expected"] <= end_date:
                e_prog += IFT_WEIGHT
            final_actual_progress.append(row["Man Hours "] * a_prog)
            final_expected_progress.append(row["Man Hours "] * e_prog)

        df["Actual_Progress_At_Final"] = final_actual_progress
        df["Expected_Progress_At_Final"] = final_expected_progress

        by_disc = df.groupby("Discipline")[["Actual_Progress_At_Final","Expected_Progress_At_Final"]].sum()
        
        if PERCENTAGE_VIEW:
            by_disc["Actual_Progress_At_Final"] = by_disc["Actual_Progress_At_Final"] / total_mh * 100
            by_disc["Expected_Progress_At_Final"] = by_disc["Expected_Progress_At_Final"] / total_mh * 100
            y_label = "Percentage of Total Works"
        else:
            y_label = "Cumulative Hours"

        st.subheader("Actual vs. Expected Works by Discipline" if PERCENTAGE_VIEW else "Actual vs. Expected Hours by Discipline")
        fig2, ax2 = plt.subplots(figsize=(8,5))
        x = range(len(by_disc.index))
        width = 0.35
        ax2.bar(
            [i - width/2 for i in x], by_disc["Actual_Progress_At_Final"],
            width=width, label='Actual Works' if PERCENTAGE_VIEW else 'Actual Hours'
        )
        ax2.bar(
            [i + width/2 for i in x], by_disc["Expected_Progress_At_Final"],
            width=width, label='Expected Works' if PERCENTAGE_VIEW else 'Expected Hours'
        )
        ax2.set_title("Actual vs. Expected Works by Discipline" if PERCENTAGE_VIEW else "Actual vs. Expected Hours by Discipline", fontsize=10)
        ax2.set_xlabel("Discipline", fontsize=9)
        ax2.set_ylabel(y_label, fontsize=9)
        ax2.set_xticks(ticks=x)
        ax2.set_xticklabels(by_disc.index, rotation=45, ha='right', fontsize=8)
        ax2.legend(fontsize=8)
        if show_grid:
            ax2.grid(True)
        plt.tight_layout()
        st.pyplot(fig2)

        # --------------------------
        # 8) PROGRESS CHARTS (STACKED BAR AND DONUT)
        # --------------------------
        total_actual = df["Actual_Progress_At_Final"].sum()
        ifr_delivered = ((df["Issued by EPC"].notna()) & (df["Issued by EPC"] <= today_date)).sum()
        total_docs = len(df)
        ifr_values = [ifr_delivered, total_docs - ifr_delivered]

        st.subheader("Progress Charts")
        
        # Stacked Bar Chart (Full Width)
        st.write("**Actual Progress Breakdown Over Time**")
        if PERCENTAGE_VIEW:
            issuance_y = [x / total_mh * 100 for x in issuance_cums]
            review_y = [x / total_mh * 100 for x in review_cums]
            final_y = [x / total_mh * 100 for x in final_cums]
            y_label_stack = "Cumulative % of Total Works"
            fmt = '%.1f'  # No % symbol
            threshold = 5.0  # Minimum percentage to show label
        else:
            issuance_y = issuance_cums
            review_y = review_cums
            final_y = final_cums
            y_label_stack = "Cumulative Man-Hours"
            fmt = '%d'
            threshold = 5.0  # Minimum man-hours to show label
        fig_stack, ax_stack = plt.subplots(figsize=(10, 6))  # Larger figure size
        ind = np.arange(len(actual_timeline))
        # Stack bars on top of each other
        bars_issuance = ax_stack.bar(ind, issuance_y, width=0.9, label='Issuance', color=stack_colors[0])
        bars_review = ax_stack.bar(ind, review_y, width=0.9, bottom=issuance_y, label='Review', color=stack_colors[1])
        bottom_for_final = [i + r for i, r in zip(issuance_y, review_y)]
        bars_final = ax_stack.bar(ind, final_y, width=0.9, bottom=bottom_for_final, label='Final Acceptance', color=stack_colors[2])
        # Add labels inside bars only for segments above threshold
        for bars, values in [(bars_issuance, issuance_y), (bars_review, review_y), (bars_final, final_y)]:
            ax_stack.bar_label(
                bars, 
                labels=[f'{v:.1f}' if v >= threshold else '' for v in values], 
                label_type='center', 
                fontsize=8, 
                color='white', 
                padding=2
            )
        n_ticks = max(1, len(ind) // 5)
        ax_stack.set_xticks(ind[::n_ticks])
        ax_stack.set_xticklabels([actual_timeline[i].strftime('%d-%b-%Y') for i in range(0, len(actual_timeline), n_ticks)], rotation=45, ha='right')
        ax_stack.set_title("Actual Progress Breakdown", fontsize=9)
        ax_stack.set_xlabel("Date", fontsize=8)
        ax_stack.set_ylabel(y_label_stack, fontsize=8)
        ax_stack.legend(fontsize=7)
        if show_grid:
            ax_stack.grid(True)
        plt.tight_layout()
        st.pyplot(fig_stack)

        # Donut Chart
        st.write("**Issued By EPC Status**")
        def ifr_autopct(pct):
            total_count = sum(ifr_values)
            docs = int(round(pct * total_count / 100.0))
            return f"{docs} docs" if docs > 0 else ""
        fig_ifr, ax_ifr = plt.subplots(figsize=(4,4))
        ax_ifr.pie(
            ifr_values, labels=["Issued by EPC", "Not Yet Issued"],
            autopct=ifr_autopct, startangle=140, wedgeprops={"width":0.4}
        )
        ax_ifr.set_title("Issued By EPC Status", fontsize=9)
        st.pyplot(fig_ifr)

        # --------------------------
        # 9) NESTED PIE CHART FOR DISCIPLINE
        # --------------------------
        st.subheader("Nested Pie Chart: Document Completion by Discipline")
        def get_doc_status(row):
            if pd.notna(row["Reply By EPC"]) and row["Flag"] == 1:
                return "Completed"
            else:
                return "Incomplete"
        df["Doc_Status"] = df.apply(get_doc_status, axis=1)
        disc_counts = df.groupby("Discipline").size()
        if disc_counts.empty:
            st.warning("No Discipline data available for pie chart.")
        else:
            fig_nested_disc, ax_nested_disc = plt.subplots(figsize=(10, 10))
            outer_labels = disc_counts.index
            outer_sizes = disc_counts.values
            n_disciplines = len(outer_labels)
            if color_scheme == "Standard":
                outer_colors = [c['color'] for c in plt.rcParamsDefault['axes.prop_cycle']][:n_disciplines]
            elif color_scheme == "Shades of Blue":
                outer_colors = ["#cce5ff", "#99ccff", "#66b2ff", "#3399ff", "#007fff"][:n_disciplines]
                if n_disciplines > 5:
                    outer_colors = sns.color_palette("Blues", n_colors=n_disciplines)
            elif color_scheme == "Shades of Green":
                outer_colors = ["#ccffcc", "#99ff99", "#66ff66", "#33cc33", "#009900"][:n_disciplines]
                if n_disciplines > 5:
                    outer_colors = sns.color_palette("Greens", n_colors=n_disciplines)
            else:
                outer_colors = sns.color_palette(seaborn_palette, n_colors=n_disciplines)
            inner_sizes = []
            inner_colors = []
            status_colors = {"Completed": "#808080", "Incomplete": "#F0F0F0"}
            for disc in disc_counts.index:
                disc_docs = df[df["Discipline"] == disc]
                for _, row in disc_docs.iterrows():
                    inner_sizes.append(1)
                    inner_colors.append(status_colors[row["Doc_Status"]])
            if not inner_sizes or sum(inner_sizes) == 0:
                st.warning("No valid data for inner pie chart (Discipline). All counts are zero or empty.")
            else:
                outer_wedges, outer_texts = ax_nested_disc.pie(
                    outer_sizes, radius=1.0, labels=None, startangle=90,
                    wedgeprops=dict(width=0.3, edgecolor='w'), colors=outer_colors
                )
                for i, (wedge, label, count) in enumerate(zip(outer_wedges, outer_labels, outer_sizes)):
                    angle = (wedge.theta2 - wedge.theta1)/2. + wedge.theta1
                    x = 1.1 * np.cos(np.deg2rad(angle))
                    y = 1.1 * np.sin(np.deg2rad(angle))
                    horizontalalignment = {-1: "right", 1: "left"}.get(np.sign(x), "center")
                    ax_nested_disc.annotate(
                        label, xy=(x, y), xytext=(1.5*np.sign(x), 0), textcoords='offset points',
                        ha=horizontalalignment, va='center', fontsize=8, fontweight='normal'
                    )
                    ax_nested_disc.annotate(
                        f"({count})", xy=(x, y), xytext=(1.5*np.sign(x), -15), textcoords='offset points',
                        ha=horizontalalignment, va='center', fontsize=10, fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.2', fc='white', alpha=0.8)
                    )
                try:
                    inner_wedges = ax_nested_disc.pie(
                        inner_sizes, radius=0.7, startangle=90, wedgeprops=dict(width=0.3, edgecolor='w'),
                        colors=inner_colors
                    )[0]
                except ValueError as e:
                    st.error(f"Error plotting inner pie chart (Discipline): {str(e)}")
                    plt.close(fig_nested_disc)
                    st.stop()
                from matplotlib.patches import Patch
                status_patches = [
                    Patch(color=status_colors["Completed"], label="Completed"),
                    Patch(color=status_colors["Incomplete"], label="Incomplete")
                ]
                ax_nested_disc.legend(
                    handles=status_patches, title="Status", loc="center left",
                    bbox_to_anchor=(1, 0.5), fontsize=8
                )
                ax_nested_disc.set_title("Documents by Discipline and Completion", fontsize=10)
                plt.tight_layout()
                st.pyplot(fig_nested_disc)

        # --------------------------
        # 10) STACKED BAR IFR/IFA/IFT BY DISCIPLINE
        # --------------------------
        df["Issued_bool"] = df["Issued by EPC"].notna().astype(int)
        df["Review_bool"] = df["Review By OE"].notna().astype(int)
        df["Reply_bool"] = df["Reply By EPC"].notna().astype(int)
        disc_counts = df.groupby("Discipline")[["Issued_bool","Review_bool","Reply_bool"]].sum()
        st.subheader("Number of Docs with Issued, Review, Reply by Discipline")
        fig4, ax4 = plt.subplots(figsize=(8,5))
        disc_counts.plot(kind="barh", stacked=True, ax=ax4)
        ax4.set_xlabel("Count of Documents", fontsize=9)
        ax4.set_ylabel("Discipline", fontsize=9)
        ax4.set_title("Document Milestone Status by Discipline", fontsize=10)
        ax4.legend(labels=["Issued", "Review", "Reply"], fontsize=8)
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)
        if show_grid:
            ax4.grid(True)
        plt.tight_layout()
        for container in ax4.containers:
            ax4.bar_label(container, label_type='center', fontsize=8)
        st.pyplot(fig4)

        # --------------------------
        # 11) DELAY BY DISCIPLINE (AS OF TODAY)
        # --------------------------
        st.subheader("Delay Percentage by Discipline (As of Today)")
        actual_prog_today = []
        expected_prog_today = []
        for _, row in df.iterrows():
            a_val = 0.0
            if pd.notna(row["Issued by EPC"]) and row["Issued by EPC"] <= today_date:
                a_val += IFR_WEIGHT
            if pd.notna(row["Review By OE"]) and row["Review By OE"] <= today_date:
                a_val += IFA_WEIGHT
            if pd.notna(row["Reply By EPC"]) and row["Reply By EPC"] <= today_date and row["Flag"] == 1:
                a_val += IFT_WEIGHT
            e_val = 0.0
            if pd.notna(row["Issuance Expected"]) and row["Issuance Expected"] <= today_date:
                e_val += IFR_WEIGHT
            if pd.notna(row["Expected review"]) and row["Expected review"] <= today_date:
                e_val += IFA_WEIGHT
            if pd.notna(row["Final Issuance Expected"]) and row["Final Issuance Expected"] <= today_date:
                e_val += IFT_WEIGHT
            actual_prog_today.append(row["Man Hours "] * a_val)
            expected_prog_today.append(row["Man Hours "] * e_val)
        df["Actual_Progress_Today"] = actual_prog_today
        df["Expected_Progress_Today"] = expected_prog_today
        disc_delay = df.groupby("Discipline")[["Actual_Progress_Today","Expected_Progress_Today"]].sum()
        disc_delay["Delay_%"] = (
            (disc_delay["Expected_Progress_Today"] - disc_delay["Actual_Progress_Today"])
            / disc_delay["Expected_Progress_Today"]
        ) * 100
        disc_delay["Delay_%"] = disc_delay["Delay_%"].fillna(0)
        fig_delay, ax_delay = plt.subplots(figsize=(8,5))
        ax_delay.bar(disc_delay.index, disc_delay["Delay_%"])
        ax_delay.set_title("Delay in % by Discipline (Today)", fontsize=10)
        ax_delay.set_xlabel("Discipline", fontsize=9)
        ax_delay.set_ylabel("Delay (%)", fontsize=9)
        ax_delay.set_xticks(range(len(disc_delay.index)))
        ax_delay.set_xticklabels(disc_delay.index, rotation=45, ha='right', fontsize=8)
        if show_grid:
            ax_delay.grid(True)
        plt.tight_layout()
        st.pyplot(fig_delay)
        st.write("Detailed Delay Data:")
        st.dataframe(disc_delay)

        # --------------------------
        # 12) FINAL MILESTONE + STATUS STACKED BAR
        # --------------------------
        df["FinalMilestone"] = df.apply(get_final_milestone, axis=1)
        st.subheader("Documents by Final Milestone (Stacked by Status)")
        group_df = (
            df.groupby(["FinalMilestone","Status"])["ID"]
              .count()
              .reset_index(name="Count")
        )
        pivoted = group_df.pivot(index="FinalMilestone", columns="Status", values="Count").fillna(0)
        order = ["NO ISSUANCE", "Issued by EPC", "Review By OE", "Reply By EPC", "Finalized"]
        pivoted = pivoted.reindex(order).dropna(how="all")

        fig_status, ax_status = plt.subplots(figsize=(7,5))
        pivoted.plot(kind="bar", stacked=True, ax=ax_status)
        for container in ax_status.containers:
            ax_status.bar_label(
                container, label_type='center', fmt='%d', fontsize=8, color='white'
            )
        ax_status.set_title("Documents by Final Milestone (Stacked by Status)", fontsize=10)
        ax_status.set_xlabel("Final Milestone", fontsize=9)
        ax_status.set_ylabel("Number of Documents", fontsize=9)
        ax_status.set_xticks(range(len(pivoted.index)))
        ax_status.set_xticklabels(pivoted.index, rotation=45, ha='right', fontsize=8)
        ax_status.legend(title="Status", fontsize=8)
        if show_grid:
            ax_status.grid(True)
        plt.tight_layout()
        st.pyplot(fig_status)

        # --------------------------
        # 13) SIMPLIFIED DELAY TABLE FOR ISSUED BY EPC
        # --------------------------
        st.subheader("Document Delays (Issued by EPC vs Expected Issuance, ≥14 Days)")
        
        # Calculate IFR delay (Issued by EPC vs Issuance Expected) and filter for ≥14 days
        today = pd.Timestamp.today().normalize()
        delays = []
        display_data = {
            "ID": [],
            "Discipline": [],
            "Document Title": [],
            "Issuance Expected": [],
            "Actual Issued": [],
            "Delay (days)": [],
            "Status": []
        }
        for i, row in df.iterrows():
            expected = row["Issuance Expected"]
            actual = row["Issued by EPC"] if pd.notna(row["Issued by EPC"]) else today
            
            if pd.isna(expected):
                delay_days = 0
            else:
                delay_days = (actual - expected).days
                # Only include positive delays ≥ 14 days
                if delay_days >= 14:
                    delays.append(delay_days)
                    display_data["ID"].append(row["ID"])
                    display_data["Discipline"].append(row["Discipline"])
                    display_data["Document Title"].append(row["Document Title"])
                    display_data["Issuance Expected"].append(
                        expected.strftime("%d-%b-%y") if pd.notna(expected) else "—"
                    )
                    display_data["Actual Issued"].append(
                        actual.strftime("%d-%b-%y") if pd.notna(actual) else "Not Issued"
                    )
                    display_data["Delay (days)"].append(delay_days)
                    display_data["Status"].append(row["Status"])
        
        df_display = pd.DataFrame(display_data)
        
        if df_display.empty:
            st.warning("No documents have an issuance delay of 14 days or more.")
        else:
            # Function to apply color formatting
            def color_delay(val):
                return 'background-color: #ffcccc'  # Light red for delays ≥14 days
            
            # Apply styling
            styler = df_display.style.applymap(color_delay, subset=['Delay (days)'])
            
            # Display styled table
            st.dataframe(styler, use_container_width=True)
            
            # Excel export with formatting using openpyxl
            def export_to_excel():
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_display.to_excel(writer, sheet_name='Delays', index=False)
                    
                    # Access the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Delays']
                    
                    # Define fill colors
                    red_fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
                    
                    # Apply formatting to delay column (column F)
                    for idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=6, max_col=6), start=2):
                        for cell in row:
                            cell.fill = red_fill
                            
                    # Apply header styling
                    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                
                return output.getvalue()
            
            excel_data = export_to_excel()
            st.download_button(
                label="Download Delays Table (Excel)",
                data=excel_data,
                file_name="document_delays.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # --------------------------
        # 14) SAVE UPDATED CSV
        # --------------------------
        df_for_export = df.copy()
        df_for_export["Issuance Expected"] = df["Issuance Expected"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notna(x) else ""
        )
        df_for_export["Expected review"] = df["Expected review"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notna(x) else ""
        )
        df_for_export["Final Issuance Expected"] = df["Final Issuance Expected"].apply(
            lambda x: x.strftime("%d-%b-%y") if pd.notna(x) else ""
        )
        st.subheader("Download Updated CSV")
        st.download_button(
            label="Download Updated CSV",
            data=df_for_export.to_csv(index=False).encode('utf-8'),
            file_name="EDDR_with_calculated_expected.csv",
            mime="text/csv"
        )

    with tab2:
        # =====================================================================================
        # TAB 2 — REVIEW TIMELINE (doc titles with status in brackets; 2-line tags; selective labeling)
        # =====================================================================================
        file_extension = CSV_INPUT_PATH.name.split('.')[-1].lower()
        if file_extension not in ['xlsx', 'xls']:
            st.info("The **Review Timeline** requires an **Excel** file with a sheet named **'Review Historical record'**.")
            st.stop()

        try:
            df_hist = pd.read_excel(CSV_INPUT_PATH, sheet_name="Review Historical record")
        except Exception:
            st.warning("Could not find a sheet named **'Review Historical record'** in the uploaded Excel file.")
            st.stop()

        orig_cols = list(df_hist.columns)
        if len(orig_cols) < 6:
            st.error("The 'Review Historical record' sheet doesn't have the expected structure (need ≥6 columns).")
            st.stop()

        base_cols = orig_cols[:4]      # ID, Discipline, Area, Document Title (expected)
        tail_cols = orig_cols[4:]      # Rev/Reviewed pairs

        # Build Rev/Reviewed pairs using ORIGINAL names, pattern-matching on normalized strings
        pairs = []
        used = set()
        for i, c in enumerate(tail_cols):
            if i in used:
                continue
            if is_rev_col(c):
                # nearest next "review-like" column
                j = i + 1
                found = False
                while j < len(tail_cols):
                    if j not in used and is_review_col(tail_cols[j]):
                        pairs.append((tail_cols[i], tail_cols[j]))
                        used.add(i); used.add(j)
                        found = True
                        break
                    j += 1
                if not found and i + 1 < len(tail_cols):
                    # adjacency fallback
                    pairs.append((tail_cols[i], tail_cols[i+1]))
                    used.add(i); used.add(i+1)

        # If still nothing, pair by position (5th with 6th, 7th with 8th, ...)
        if not pairs:
            for k in range(0, len(tail_cols), 2):
                left = tail_cols[k]
                right = tail_cols[k+1] if k+1 < len(tail_cols) else None
                if right is not None:
                    pairs.append((left, right))

        if not pairs:
            st.error("No valid (RevX, Review/Reviewed) pairs detected.")
            st.write("Columns from 5th onward:", tail_cols)
            st.stop()

        with st.expander("Detected column pairs", expanded=False):
            st.write(pairs)

        # Parse dates in all Rev/Review columns
        for rev_c, revw_c in pairs:
            df_hist[rev_c] = df_hist[rev_c].apply(robust_parse_date)
            df_hist[revw_c] = df_hist[revw_c].apply(robust_parse_date)

        # Filter rows where first Rev column (e.g., Rev0) is not null
        first_rev_col = pairs[0][0]
        df_sel = df_hist[df_hist[first_rev_col].notna()].copy()
        if df_sel.empty:
            st.warning(f"No rows have a non-null **{first_rev_col}** (initial submission). Nothing to plot.")
            st.stop()

        # Build display label: use **Document Title only** on the y-axis to make space
        title_candidates = ["Document Title", "Title", base_cols[min(3, len(base_cols)-1)]]
        title_col = next((c for c in title_candidates if c in df_sel.columns), title_candidates[-1])

        # Select by Title
        st.subheader("Select Documents to Plot")
        choices = st.multiselect(
            "Choose one or more documents (must have initial submission date).",
            options=sorted(df_sel[title_col].astype(str).unique().tolist())
        )
        if not choices:
            st.info("Select at least one document title to render the timeline.")
            st.stop()

        df_plot = df_sel[df_sel[title_col].astype(str).isin(choices)].copy()

        # Build actual segments (title, rev_tag, submit, review) for ALL pairs
        actual_segments = []
        for _, r in df_plot.iterrows():
            doc_title = str(r.get(title_col, ""))
            for rev_c, revw_c in pairs:
                submit_dt = r.get(rev_c, pd.NaT)
                review_dt = r.get(revw_c, pd.NaT)
                if pd.notna(submit_dt):
                    m = re.findall(r'\d+', normalize_header(rev_c))
                    rev_tag = f"Rev{m[0]}" if m else normalize_header(rev_c)
                    actual_segments.append({
                        "title": doc_title,
                        "rev": rev_tag,
                        "submit": submit_dt,
                        "review": review_dt if pd.notna(review_dt) else None
                    })

        if not actual_segments:
            st.warning("No valid submission dates found to plot.")
            st.stop()

        # Build expected segments from first sheet (Issuance Expected, Expected review, Final Issuance Expected)
        expected_segments = []
        df_expected = df[df["Document Title"].astype(str).isin(choices)].drop_duplicates(subset=["Document Title"])  # Ensure unique titles
        for _, r in df_expected.iterrows():
            doc_title = str(r["Document Title"])
            ifr_exp = robust_parse_date(r["Issuance Expected"])
            ifa_exp = robust_parse_date(r["Expected review"])
            ift_exp = robust_parse_date(r["Final Issuance Expected"])
            if pd.notna(ifr_exp):  # Only include if Issuance Expected is not null
                expected_segments.append({
                    "title": doc_title,
                    "ifr_exp": ifr_exp,
                    "ifa_exp": ifa_exp if pd.notna(ifa_exp) else None,
                    "ift_exp": ift_exp if pd.notna(ift_exp) else None
                })

        if not expected_segments:
            st.warning("No valid expected dates found for selected documents in the first sheet.")
            # Proceed with actual segments only

        # y-axis (one row per document title with status in brackets)
        status_map = df.set_index("Document Title")["Status"].to_dict()
        titles = sorted(set(s["title"] for s in actual_segments))
        y_positions = {t: i for i, t in enumerate(titles)}
        # Append status to titles for y-axis labels
        title_labels = [f"{t} [{status_map.get(t, 'Unknown')}]" for t in titles]

        # Identify first and last points to label for each document
        label_points = {}
        for title in titles:
            segs = [s for s in actual_segments if s["title"] == title]
            if not segs:
                continue
            # First submission: earliest submit date
            first_seg = min(segs, key=lambda s: s["submit"])
            label_points[(title, first_seg["submit"], "submit")] = first_seg["rev"]
            # Last point: latest submit or review date
            dates = [(s["submit"], "submit", s["rev"]) for s in segs]
            dates.extend([(s["review"], "review", s["rev"]) for s in segs if s["review"] is not None])
            if dates:
                last_date, last_type, last_rev = max(dates, key=lambda x: x[0])
                label_points[(title, last_date, last_type)] = last_rev

        # Debug: Show which points will be labeled
        with st.expander("Points Selected for Labeling", expanded=False):
            st.write(label_points)

        # Plot — two-line labels with above/below placement and selective labeling
        st.subheader("Review Timeline (Submission ➜ Review with Expected Dates)")
        fig_t, ax_t = plt.subplots(figsize=(12, 1.1*max(4, len(titles))))  # Increased height for more labels
        ax_t.xaxis_date()  # Set x-axis to datetime immediately
        ax_t.xaxis.set_major_formatter(mdates.DateFormatter("%d-%b-%Y"))
        ax_t.xaxis.set_major_locator(mdates.AutoDateLocator())
        submit_marker = 'o'
        review_marker = 's'   # square
        expected_marker = '^'  # triangle for expected dates
        label_offset_y = 12   # Initial vertical offset (above or below the line, pixels)
        font_size = 7         # Slightly larger font for readability
        max_jitter_x = 20     # Maximum horizontal jitter (pixels)
        vertical_step = 10    # Vertical offset step for stacking (pixels)
        label_width_days = 3  # Tighter overlap detection
        expected_color = '#ff7f0e'  # Match S-Curve expected color

        color_cycle = plt.rcParams['axes.prop_cycle'].by_key().get('color', ['#1f77b4'])
        title_color_map = {t: color_cycle[i % len(color_cycle)] for i, t in enumerate(titles)}

        # Track occupied label regions separately for above and below
        occupied_regions_above = []
        occupied_regions_below = []

        def is_overlapping(x_center, y_center, x_min, x_max, y_min, y_max, is_actual=False):
            """Check if a new label overlaps with existing labels in the same group (above or below)."""
            regions = occupied_regions_above if is_actual else occupied_regions_below
            for region in regions:
                ox_center, oy_center, ox_min, ox_max, oy_min, oy_max = region
                if (x_max > ox_min and x_min < ox_max and
                    y_max > oy_min and y_min < oy_max):
                    return True
            return False

        def get_label_position(x, y, title, ts, is_actual=False):
            """Calculate label position: above for actual, below for expected, with stacking."""
            base_y_offset = label_offset_y if is_actual else -label_offset_y
            y_offset = base_y_offset
            x_jitter = 0
            attempt = 0
            max_attempts = 10  # Limit stacking to prevent excessive spread

            # Convert x (datetime) to numeric for collision detection
            x_num = mdates.date2num(x)
            x_min = x_num - label_width_days / 2
            x_max = x_num + label_width_days / 2
            y_min = y + (y_offset - 5) / 100  # Approximate height in y-units
            y_max = y + (y_offset + 15) / 100

            while is_overlapping(x_num, y + y_offset / 100, x_min, x_max, y_min, y_max, is_actual):
                attempt += 1
                if attempt % 2 == 0:
                    # Vertical stacking (up for actual, down for expected)
                    y_offset += vertical_step if is_actual else -vertical_step
                else:
                    # Horizontal jitter (alternate left/right)
                    x_jitter = (-1) ** attempt * (attempt // 2 + 1) * 10
                    if abs(x_jitter) > max_jitter_x:
                        x_jitter = 0
                        y_offset += vertical_step if is_actual else -vertical_step
                y_min = y + (y_offset - 5) / 100
                y_max = y + (y_offset + 15) / 100
                if attempt >= max_attempts:
                    break  # Accept slight overlap if necessary

            (occupied_regions_above if is_actual else occupied_regions_below).append(
                (x_num, y + y_offset / 100, x_min, x_max, y_min, y_max)
            )
            return x_jitter, y_offset

        # Plot actual segments (submission and review)
        actual_segments.sort(key=lambda z: (y_positions[z["title"]], z["submit"], z["review"] or z["submit"]))
        for seg in actual_segments:
            y = y_positions[seg["title"]]
            c = title_color_map[seg["title"]]
            x0 = seg["submit"]
            x1 = seg["review"]

            # Plot submission marker
            ax_t.plot([x0], [y], marker=submit_marker, markersize=7, color=c, linestyle='None')
            # Label only if it's the first submission
            if (seg["title"], x0, "submit") in label_points:
                x_jitter0, y_offset0 = get_label_position(x0, y, seg["title"], x0, is_actual=True)
                ax_t.annotate(f'{seg["rev"]}\n{x0.strftime("%d-%b-%y")}',
                              xy=(x0, y), xytext=(x_jitter0, y_offset0),
                              textcoords='offset points', ha='center', va='bottom',
                              fontsize=font_size, bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="none", alpha=0.85))

            # Plot review (if any)
            if x1 is not None:
                ax_t.plot([x0, x1], [y, y], color=c, linewidth=2, alpha=0.9)
                ax_t.plot([x1], [y], marker=review_marker, markersize=6, color=c, linestyle='None')
                # Label only if it's the last point
                if (seg["title"], x1, "review") in label_points:
                    x_jitter1, y_offset1 = get_label_position(x1, y, seg["title"], x1, is_actual=True)
                    ax_t.annotate(f'{seg["rev"]} review\n{x1.strftime("%d-%b-%y")}',
                                  xy=(x1, y), xytext=(x_jitter1, y_offset1),
                                  textcoords='offset points', ha='center', va='bottom',
                                  fontsize=font_size, bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="none", alpha=0.85))
            else:
                # No review yet: short tick to indicate in-progress
                ax_t.plot([x0, x0 + pd.Timedelta(days=1)], [y, y], color=c, linewidth=1.5, alpha=0.6)

        # Plot expected segments (IFR Exp, IFA Exp, IFT Exp)
        for seg in expected_segments:
            y = y_positions.get(seg["title"])
            if y is None:
                continue  # Skip if title not in selected documents
            dates = []
            labels = []
            if pd.notna(seg["ifr_exp"]):
                dates.append(seg["ifr_exp"])
                labels.append("Submission")
            if seg["ifa_exp"] is not None:
                dates.append(seg["ifa_exp"])
                labels.append("Review")
            if seg["ift_exp"] is not None:
                dates.append(seg["ift_exp"])
                labels.append("Final Doc")

            if dates:
                # Ensure dates are pd.Timestamp, re-parse if strings
                valid_dates = []
                valid_labels = []
                for d, lbl in zip(dates, labels):
                    if isinstance(d, str):
                        d = robust_parse_date(d)
                    if pd.notna(d):
                        valid_dates.append(d.to_pydatetime() if isinstance(d, pd.Timestamp) else d)
                        valid_labels.append(lbl)
                if valid_dates:
                    # Sort dates to ensure correct plotting order
                    date_label_pairs = sorted(zip(valid_dates, valid_labels), key=lambda x: x[0])
                    sorted_dates, sorted_labels = zip(*date_label_pairs) if date_label_pairs else ([], [])
                    sorted_dates = list(sorted_dates)  # Convert to list of datetime.datetime
                    if sorted_dates:
                        # Plot dotted line connecting expected dates
                        ax_t.plot(sorted_dates, [y] * len(sorted_dates), linestyle=':', color=expected_color, linewidth=2, alpha=0.7, label="Expected Timeline" if y == y_positions[titles[0]] else "")
                        # Plot markers for expected dates
                        for x, label in zip(sorted_dates, sorted_labels):
                            ax_t.plot([x], [y], marker=expected_marker, markersize=6, color=expected_color, linestyle='None')
                            x_jitter, y_offset = get_label_position(x, y, seg["title"], x, is_actual=False)
                            ax_t.annotate(f'{label}\n{x.strftime("%d-%b-%y")}',
                                          xy=(x, y), xytext=(x_jitter, y_offset),
                                          textcoords='offset points', ha='center', va='top',
                                          fontsize=font_size, bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="none", alpha=0.85))

        from matplotlib.lines import Line2D
        
        # Legend entries for markers only (dot = submission, square = review)
        extra_legend = [
            Line2D([0], [0], marker=submit_marker, linestyle='None', color='none',
                   markerfacecolor='#1f77b4', markeredgecolor='#1f77b4', markersize=7, label='Submission'),
            Line2D([0], [0], marker=review_marker, linestyle='None', color='none',
                   markerfacecolor='#1f77b4', markeredgecolor='#1f77b4', markersize=7, label='Review'),
        ]

        ax_t.set_yticks([y_positions[t] for t in titles])
        ax_t.set_yticklabels(title_labels, fontsize=8)
        ax_t.set_ylim(-0.6, len(titles) - 0.4)
        ax_t.set_xlabel("Date", fontsize=9)
        ax_t.set_title("Submission → Review Timeline with Expected Dates", fontsize=11)
        if show_grid:
            ax_t.grid(True, axis='x', linestyle='--', alpha=0.35)
        handles, labels = ax_t.get_legend_handles_labels()
        ax_t.legend(handles=extra_legend + handles, fontsize=8, loc='upper left')

        plt.xticks(rotation=45)
        plt.tight_layout()
        st.pyplot(fig_t)

        # Compact table of plotted items (actual + expected)
        st.markdown("**Plotted Revisions and Expected Dates (compact table)**")
        tbl_rows = []
        for s in actual_segments:
            tbl_rows.append({
                "Document Title": s["title"],
                "Type": "Actual",
                "Revision": s["rev"],
                "Submission": s["submit"].strftime("%d-%b-%y") if pd.notna(s["submit"]) else "—",
                "Review": s["review"].strftime("%d-%b-%y") if s["review"] else "—",
                "IFR Exp": "—",
                "IFA Exp": "—",
                "IFT Exp": "—"
            })
        for s in expected_segments:
            y = y_positions.get(s["title"])
            if y is None:
                continue
            tbl_rows.append({
                "Document Title": s["title"],
                "Type": "Expected",
                "Revision": "—",
                "Submission": "—",
                "Review": "—",
                "First Submission": s["ifr_exp"].strftime("%d-%b-%y") if pd.notna(s["ifr_exp"]) else "—",
                "Document Review": s["ifa_exp"].strftime("%d-%b-%y") if s["ifa_exp"] else "—",
                "Final Submission": s["ift_exp"].strftime("%d-%b-%y") if s["ift_exp"] else "—"
            })
        st.dataframe(pd.DataFrame(tbl_rows))

if __name__ == "__main__":
    main()
